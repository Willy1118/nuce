# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os, glob, subprocess, sys
from datetime import datetime

# ---------- Excel 初始化 ----------
def init_excel():
    if not os.path.exists("user.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"
        ws.append(["Username", "Password", "Chips", "CardID", "CardPassword", "Money"])
        wb.save("user.xlsx")

    if not os.path.exists("game_log.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(["user_name", "Time", "Game", "payout_ratio", "Chips", "WinOrLose", "Outcome"])
        wb.save("game_log.xlsx")

# ---------- 資料處理 ----------
def find_user(username):
    wb = load_workbook("user.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == username:
            return row, wb
    return None, wb

def register_user(username, password, card_id, card_pwd):
    row, wb = find_user(username)
    if row: return False
    ws = wb.active
    ws.append([username, password, 0, card_id, card_pwd, 10000])
    wb.save("user.xlsx")
    return True

def validate_user(username, password):
    row, _ = find_user(username)
    return row and row[1].value == password

def get_chips(username):
    row, _ = find_user(username)
    return row[2].value if row else 0

def recharge_with_card(username, card_id, password, amount):
    try: amount = float(amount)
    except: return False, "金額格式錯誤"
    row, wb = find_user(username)
    if not row or row[3].value != card_id or row[4].value != password:
        return False, "卡號或卡密錯誤"
    if row[5].value < amount:
        return False, "存款不足"
    row[5].value -= amount
    row[2].value += amount
    wb.save("user.xlsx")
    return True, "儲值成功"

def record_game(username, game, magnification, chips, result):
    wb = load_workbook("game_log.xlsx")
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payout = 1 if result == "win" else 0 if result == "push" else -1
    outcome = chips * payout if result != "lose" else -chips
    ws.append([username, now, game, magnification, chips, result, outcome])
    wb.save("game_log.xlsx")

def read_records(username):
    records = []
    if not os.path.exists("game_log.xlsx"):
        return records
    wb = load_workbook("game_log.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            records.append((row[1], row[2], row[0], row[3], row[4], row[5], row[6]))
    return records

# ---------- Tkinter 樣式 ----------
def apply_style(root):
    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TButton", background="#1c1c1c", foreground="#FFD700", font=("Arial", 20, "bold"))
    style.map("TButton", background=[("active", "#FFD700")], foreground=[("active", "#1c1c1c")])
    style.configure("TLabel", background="#1c1c1c", foreground="#FFD700", font=("Arial", 20))
    style.configure("TEntry", fieldbackground="#333", foreground="#FFD700", background="#1c1c1c", font=("Arial", 20))
    style.configure("TCombobox", fieldbackground="#333", background="#1c1c1c", foreground="#FFD700", font=("Arial", 20))
    style.configure("Treeview", background="#2e2e2e", fieldbackground="#2e2e2e", foreground="#FFD700", font=("Arial", 18))
    style.configure("Treeview.Heading", background="#1c1c1c", foreground="#FFD700", font=("Arial", 20, "bold"))

# ---------- 主畫面 ----------
def open_main_window(username):
    window = tk.Tk()
    apply_style(window)
    window.geometry("1200x700+0+0")
    window.configure(bg="#1c1c1c")

    chips_var = tk.StringVar()
    time_var = tk.StringVar()

    def update_time():
        time_var.set("🕒 現在時間：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        chips_var.set(f"💰 籌碼：{get_chips(username)} 元")
        window.after(1000, update_time)

    left_frame = tk.Frame(window, bg="#1c1c1c")
    center_frame = tk.Frame(window, bg="#1c1c1c")
    right_frame = tk.Frame(window, bg="#1c1c1c")

    left_frame.place(x=0, y=150, width=400, height=650)
    center_frame.place(x=400, y=150, width=400, height=650)
    right_frame.place(x=800, y=100, width=400, height=650)

    ttk.Label(left_frame, text=f"🎰 玩家：{username}").pack(pady=20)
    ttk.Label(left_frame, textvariable=time_var).pack(pady=20)
    ttk.Label(left_frame, textvariable=chips_var).pack(pady=20)
    ttk.Button(left_frame, text="查詢紀錄", command=lambda: show_records(window, username)).pack(pady=20)

    ttk.Label(center_frame, text="🎮 選擇遊戲：").pack(pady=20)
    game_files = glob.glob("game/*.py")
    game_list = [os.path.splitext(os.path.basename(f))[0] for f in game_files]
    selected_game = tk.StringVar()
    dropdown = ttk.Combobox(center_frame, values=game_list, textvariable=selected_game, width=20, font=("Arial", 20))
    dropdown.pack(pady=30)
    window.option_add("*TCombobox*Listbox.font", ("Arial", 20))

    def launch_game():
        game = selected_game.get()
        if game:
            window.destroy()
            path = os.path.join("game", game + ".py")
            subprocess.Popen([sys.executable, path, username])

    ttk.Button(center_frame, text="開始遊戲", command=launch_game).pack(pady=30)

    ttk.Label(right_frame, text="💳 加值區塊").pack(pady=20)
    ttk.Label(right_frame, text="卡號（CardID）").pack()
    card_id_entry = ttk.Entry(right_frame, font=("Arial", 20))
    card_id_entry.pack(pady=10)

    ttk.Label(right_frame, text="卡密").pack()
    pwd_entry = ttk.Entry(right_frame, show="*", font=("Arial", 20))
    pwd_entry.pack(pady=10)

    ttk.Label(right_frame, text="金額").pack()
    amount_entry = ttk.Entry(right_frame, font=("Arial", 20))
    amount_entry.pack(pady=10)

    def do_recharge():
        card_id = card_id_entry.get()
        password = pwd_entry.get()
        amount = amount_entry.get()
        success, msg = recharge_with_card(username, card_id, password, amount)
        messagebox.showinfo("訊息", msg)
        update_time()

    ttk.Button(right_frame, text="確認加值", command=do_recharge).pack(pady=30)

    update_time()
    window.mainloop()

def show_records(window, username):
    top = tk.Toplevel(window)
    top.title("歷史紀錄")
    top.geometry("1200x400")

    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 12), rowheight=35)
    style.configure("Treeview.Heading", font=("Arial", 14, "bold"))

    tree = ttk.Treeview(top, columns=("遊戲時間", "遊戲", "玩家", "輸贏賠率", "該注籌碼", "輸贏結果", "輸贏金額"), show="headings")
    for col in tree["columns"]:
        tree.heading(col, text=col)
        tree.column(col, width=150)

    for r in read_records(username):
        tree.insert("", "end", values=r)

    tree.pack(fill="both", expand=True)

# ---------- 登入與註冊 ----------
def open_start_window():
    window = tk.Tk()
    apply_style(window)
    window.geometry("1200x700+0+0")
    window.configure(bg="#1c1c1c")

    time_var = tk.StringVar()

    def update_time():
        time_var.set("🕒 現在時間：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        window.after(1000, update_time)

    frame = ttk.Frame(window)
    frame.pack(pady=100)

    ttk.Label(window, textvariable=time_var, font=("Arial", 20)).pack(pady=20)
    btn_login = ttk.Button(frame, text="登入", width=20)
    btn_register = ttk.Button(frame, text="註冊", width=20)
    btn_login.grid(row=0, column=0, pady=10)
    btn_register.grid(row=1, column=0, pady=10)

    btn_multi = ttk.Button(frame, text="多人登入遊戲", width=30)
    btn_multi.grid(row=2, column=0, pady=5)
    btn_multi.config(command=lambda: [window.destroy(), subprocess.Popen([sys.executable, "perudo.py"])])

    def switch(mode):
        for widget in frame.winfo_children():
            widget.destroy()

        ttk.Label(frame, text="帳號：").grid(row=0, column=0, sticky="e", pady=10)
        user_entry = ttk.Entry(frame, font=("Arial", 20))
        user_entry.grid(row=0, column=1, pady=10)

        ttk.Label(frame, text="密碼：").grid(row=1, column=0, sticky="e", pady=10)
        pass_entry = ttk.Entry(frame, show="*", font=("Arial", 20))
        pass_entry.grid(row=1, column=1, pady=10)

        if mode == "register":
            ttk.Label(frame, text="卡帳：").grid(row=2, column=0, sticky="e", pady=10)
            card_entry = ttk.Entry(frame, font=("Arial", 20))
            card_entry.grid(row=2, column=1, pady=10)

            ttk.Label(frame, text="卡密：").grid(row=3, column=0, sticky="e", pady=10)
            card_pwd_entry = ttk.Entry(frame, show="*", font=("Arial", 20))
            card_pwd_entry.grid(row=3, column=1, pady=10)

        def submit():
            u, p = user_entry.get(), pass_entry.get()
            if mode == "login":
                if validate_user(u, p):
                    window.destroy()
                    open_main_window(u)
                else:
                    messagebox.showerror("錯誤", "帳號或密碼錯誤")
            else:
                card_id = card_entry.get()
                card_pwd = card_pwd_entry.get()
                if register_user(u, p, card_id, card_pwd):
                    messagebox.showinfo("成功", "註冊成功，請重新登入")
                    switch("login")
                else:
                    messagebox.showerror("錯誤", "帳號已存在")

        row_idx = 4 if mode == "register" else 2
        ttk.Button(frame, text="確認", command=submit).grid(row=row_idx, column=0, columnspan=2, pady=20)

    btn_login.config(command=lambda: switch("login"))
    btn_register.config(command=lambda: switch("register"))

    update_time()
    window.mainloop()

# ---------- 程式入口 ----------
if __name__ == "__main__":
    init_excel()
    args = sys.argv

    if len(args) >= 3 and args[1] == "--back":
        username = args[2]
        open_main_window(username)
    else:
        open_start_window()